"""File loading: ZIP extraction + streaming Excel/CSV to SQLite."""

from __future__ import annotations

import csv
import io
import os
import sqlite3
import zipfile
from typing import Any, Iterator

import openpyxl

from shared.db import (
    safe_table_name,
    register_table,
    store_table_streaming,
    table_exists,
    table_row_count,
    read_table_columns,
    quote_id,
)
from shared.utils import make_unique

RAW_META_PREVIEW_ROWS = int(os.getenv("RAW_META_PREVIEW_ROWS", "20"))

_META_COLUMNS = ["FILE_NAME", "RECORD_ID"]


def _file_name_from_key(table_key: str) -> str:
    """Extract human-readable file name from a table_key like 'path/to/file.xlsx::Sheet1'."""
    path_part = table_key.split("::")[0]
    return path_part.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]


def _clean_header(cells: list) -> list[str]:
    """Clean a header row: strip, uppercase, fill blanks."""
    result = []
    for i, c in enumerate(cells):
        if c is None or str(c).strip() == "":
            result.append(f"Column_{i + 1}")
        else:
            result.append(str(c).strip().upper())
    return make_unique(result)


def _clean_cell(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip().upper()
    return s if s else None


def _is_empty_row(vals: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in vals)


def _iter_csv_rows(text: str) -> Iterator[list[str]]:
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        yield list(row)


def _pad_row(row: list, max_cols: int) -> list:
    """Pad or truncate a row to exactly max_cols."""
    if len(row) < max_cols:
        return list(row) + [None] * (max_cols - len(row))
    if len(row) > max_cols:
        return list(row)[:max_cols]
    return list(row)


def _load_excel_sheet(
    conn: sqlite3.Connection,
    table_key: str,
    wb: openpyxl.Workbook,
    sheet_name: str,
    warnings: list[dict],
    commit: bool = True,
) -> None:
    """Load a single Excel sheet using an already-open workbook (single-pass)."""
    try:
        ws = wb[sheet_name]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not all_rows:
            return

        header_raw = all_rows[0]
        max_cols = max(len(r) for r in all_rows)
        if max_cols <= 0:
            max_cols = len(header_raw) if header_raw else 1

        raw_cols = [f"RAW_{i + 1}" for i in range(max_cols)]
        raw_name = safe_table_name("raw", table_key)
        store_table_streaming(
            conn, raw_name, raw_cols,
            (_pad_row(r, max_cols) for r in all_rows),
            commit=commit,
        )

        base_header = _clean_header(header_raw)
        file_name = _file_name_from_key(table_key)
        header = _META_COLUMNS + base_header
        num_base = len(base_header)

        def _clean_gen():
            record_id = 0
            for r in all_rows[1:]:
                if _is_empty_row(r):
                    continue
                record_id += 1
                yield [file_name, str(record_id)] + [
                    _clean_cell(r[i]) if i < len(r) else None
                    for i in range(num_base)
                ]

        tbl_name = safe_table_name("tbl", table_key)
        store_table_streaming(conn, tbl_name, header, _clean_gen(), commit=commit)
        register_table(conn, table_key, tbl_name, commit=commit)
    except Exception as e:
        warnings.append({"file": table_key, "message": str(e)})


def _load_csv(
    conn: sqlite3.Connection,
    table_key: str,
    csv_bytes: bytes,
    warnings: list[dict],
    commit: bool = True,
) -> None:
    """Load a CSV file using a single parse pass."""
    try:
        text = csv_bytes.decode("utf-8", errors="replace")
        all_rows = list(_iter_csv_rows(text))
        if not all_rows:
            return

        header_raw = all_rows[0]
        max_cols = max(len(r) for r in all_rows)
        if max_cols <= 0:
            max_cols = len(header_raw) if header_raw else 1

        raw_cols = [f"RAW_{i + 1}" for i in range(max_cols)]
        raw_name = safe_table_name("raw", table_key)
        store_table_streaming(
            conn, raw_name, raw_cols,
            (_pad_row(r, max_cols) for r in all_rows),
            commit=commit,
        )

        base_header = _clean_header(header_raw)
        file_name = _file_name_from_key(table_key)
        header = _META_COLUMNS + base_header
        num_base = len(base_header)

        def _clean_gen():
            record_id = 0
            for r in all_rows[1:]:
                if _is_empty_row(r):
                    continue
                record_id += 1
                yield [file_name, str(record_id)] + [
                    _clean_cell(r[i]) if i < len(r) else None
                    for i in range(num_base)
                ]

        tbl_name = safe_table_name("tbl", table_key)
        store_table_streaming(conn, tbl_name, header, _clean_gen(), commit=commit)
        register_table(conn, table_key, tbl_name, commit=commit)
    except Exception as e:
        warnings.append({"file": table_key, "message": str(e)})


def load_zip_to_session(conn: sqlite3.Connection, file_data: bytes) -> tuple[dict[str, list], list[dict]]:
    """Parse a ZIP archive and stream all files into the session SQLite.

    Opens each Excel workbook exactly once and parses each file in a single
    pass.  All DB writes are batched into one commit at the end.

    Returns ({}, warnings_list).  The first element is kept for interface
    compatibility but is always empty (raw data lives in raw__* tables).
    """
    warnings: list[dict] = []

    with zipfile.ZipFile(io.BytesIO(file_data)) as zf:
        for entry in zf.infolist():
            if entry.is_dir():
                continue
            name = entry.filename
            lower = name.lower()

            if lower.endswith((".xlsx", ".xlsm")):
                data = zf.read(name)
                wb = openpyxl.load_workbook(
                    io.BytesIO(data), read_only=True, data_only=True,
                )
                try:
                    for sheet in wb.sheetnames:
                        key = f"{name}::{sheet}"
                        _load_excel_sheet(
                            conn, key, wb, sheet, warnings, commit=False,
                        )
                finally:
                    wb.close()

            elif lower.endswith(".csv"):
                key = f"{name}::"
                _load_csv(conn, key, zf.read(name), warnings, commit=False)

    conn.commit()
    return {}, warnings


def get_raw_array_from_table(
    conn: sqlite3.Connection,
    table_key: str,
    limit: int | None = None,
) -> list[list[Any]]:
    raw_name = safe_table_name("raw", table_key)
    if not table_exists(conn, raw_name):
        return []
    cols = read_table_columns(conn, raw_name)
    if not cols:
        return []
    select_cols = ", ".join(quote_id(c) for c in cols)
    sql = f"SELECT {select_cols} FROM {quote_id(raw_name)}"
    params: tuple[Any, ...] = ()
    if limit is not None:
        sql += " LIMIT ?"
        params = (int(limit),)
    rows = conn.execute(sql, params).fetchall()
    return [[r[c] for c in cols] for r in rows]


def _build_columns_from_header(
    header_row: list[Any],
    custom_column_names: dict[int, str] | None = None,
) -> list[str]:
    columns: list[str] = []
    for i, cell in enumerate(header_row):
        if custom_column_names and i in custom_column_names:
            columns.append(str(custom_column_names[i]))
        elif cell is None or str(cell).strip() == "":
            columns.append(f"Column_{i + 1}")
        else:
            columns.append(str(cell).strip())
    normalized = [
        str(c).strip().upper() if str(c).strip() else f"COLUMN_{i + 1}"
        for i, c in enumerate(columns)
    ]
    return make_unique(normalized)


def rebuild_table_from_raw_table(
    conn: sqlite3.Connection,
    table_key: str,
    header_row_index: int,
    custom_column_names: dict[int, str] | None = None,
) -> None:
    raw_name = safe_table_name("raw", table_key)
    if not table_exists(conn, raw_name):
        raise ValueError("Raw table data not found for this table. Please re-upload.")

    raw_cols = read_table_columns(conn, raw_name)
    if not raw_cols:
        raise ValueError("Raw table has no columns.")

    total_rows = table_row_count(conn, raw_name)
    if header_row_index < 0 or header_row_index >= total_rows:
        raise ValueError(f"headerRowIndex {header_row_index} is out of range.")

    select_cols = ", ".join(quote_id(c) for c in raw_cols)
    header_row = conn.execute(
        f"SELECT {select_cols} FROM {quote_id(raw_name)} WHERE rowid = ?",
        (header_row_index + 1,),
    ).fetchone()
    if not header_row:
        raise ValueError("Header row not found in raw table.")

    header_values = [header_row[c] for c in raw_cols]
    final_columns = _build_columns_from_header(header_values, custom_column_names)
    if not final_columns:
        raise ValueError("Could not infer any columns from selected header row.")

    # Pass 1: detect columns that are entirely empty below the selected header row.
    has_value = [False] * len(final_columns)
    row_cursor = conn.execute(
        f"SELECT {select_cols} FROM {quote_id(raw_name)} WHERE rowid > ?",
        (header_row_index + 1,),
    )
    for row in row_cursor:
        for i in range(len(final_columns)):
            raw_val = row[raw_cols[i]] if i < len(raw_cols) else None
            if raw_val is not None and str(raw_val).strip() != "":
                has_value[i] = True

    valid_idx = [i for i, ok in enumerate(has_value) if ok]
    if not valid_idx:
        valid_idx = list(range(len(final_columns)))
    base_output_columns = [final_columns[i] for i in valid_idx]
    file_name = _file_name_from_key(table_key)
    output_columns = _META_COLUMNS + base_output_columns

    def _data_gen():
        record_id = 0
        cur = conn.execute(
            f"SELECT {select_cols} FROM {quote_id(raw_name)} WHERE rowid > ?",
            (header_row_index + 1,),
        )
        for row in cur:
            out_row: list[str | None] = []
            non_empty = False
            for i in valid_idx:
                raw_val = row[raw_cols[i]] if i < len(raw_cols) else None
                cleaned = _clean_cell(raw_val)
                out_row.append(cleaned)
                if cleaned is not None and cleaned != "":
                    non_empty = True
            if non_empty:
                record_id += 1
                yield [file_name, str(record_id)] + out_row

    tbl_name = safe_table_name("tbl", table_key)
    store_table_streaming(conn, tbl_name, output_columns, _data_gen())
    register_table(conn, table_key, tbl_name)


def array_to_objects(
    raw_arr: list[list],
    header_row_index: int,
    custom_column_names: dict[int, str] | None = None,
) -> list[dict]:
    """Convert a 2D raw array into a list of row dicts, given a header row index."""
    if not raw_arr or header_row_index >= len(raw_arr):
        return []

    header_row = raw_arr[header_row_index]
    data_rows = raw_arr[header_row_index + 1:]

    columns: list[str] = []
    for i, cell in enumerate(header_row):
        if custom_column_names and i in custom_column_names:
            columns.append(custom_column_names[i])
        elif cell is None or str(cell).strip() == "":
            columns.append(f"Column_{i + 1}")
        else:
            columns.append(str(cell).strip())

    return [
        {columns[i]: (row[i] if i < len(row) else None) for i in range(len(columns))}
        for row in data_rows
    ]


def clean_rows_sql(rows: list[dict]) -> list[dict]:
    """Clean a list of row-dicts: trim+uppercase keys and values, drop empty rows/cols."""
    if not rows:
        return []

    cleaned = []
    for row in rows:
        new_row = {}
        for k, v in row.items():
            clean_key = str(k).strip().upper()
            if isinstance(v, str):
                new_row[clean_key] = v.strip().upper()
            else:
                new_row[clean_key] = v
        if any(val is not None and val != "" for val in new_row.values()):
            cleaned.append(new_row)

    if not cleaned:
        return []

    all_cols = list(cleaned[0].keys())
    valid_cols = [
        c for c in all_cols
        if any(row.get(c) is not None and row.get(c) != "" for row in cleaned)
    ]

    return [{c: row.get(c) for c in valid_cols} for row in cleaned]
