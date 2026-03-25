"""Header normalisation routes: per-table schema mapping + procurement mapping."""

from __future__ import annotations

import importlib.util
import io
import os
import sys

from flask import Blueprint, jsonify, request, send_file

from shared.db import get_session_db, lookup_sql_name, read_table, read_table_columns, table_exists, table_row_count
from routes.summary_routes import execute_operation_kernel, _session_lock


def _load_mod(name: str, path: str):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)  # type: ignore[arg-type]
    sys.modules[name] = mod
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


_hn_dir = os.path.join(os.path.dirname(__file__), "..", "header-normalisation")

if _hn_dir not in sys.path:
    sys.path.insert(0, _hn_dir)

_hn_service = _load_mod("hn_service", os.path.join(_hn_dir, "service.py"))

get_table_preview = _hn_service.get_table_preview

header_normalisation_bp = Blueprint("header_normalisation_bp", __name__)


# ---------------------------------------------------------------------------
# Per-table header normalisation (new step between Inventory and Append)
# ---------------------------------------------------------------------------

@header_normalisation_bp.route("/header-norm-run", methods=["POST"])
def header_norm_run():
    try:
        body = request.get_json(force=True, silent=True) or {}
        payload, status = execute_operation_kernel(
            session_id=body.get("sessionId"),
            operation="header_norm_run",
            api_key=body.get("apiKey"),
            input_data={},
            options={"mode": "pipeline", "autoPrepare": True, "persist": True},
            request_id=request.headers.get("X-Request-Id"),
        )
        if status != 200:
            return jsonify({"error": payload.get("error"), "missing_requirements": payload.get("missing_requirements")}), status
        return jsonify(payload.get("result") or {})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@header_normalisation_bp.route("/header-norm-apply", methods=["POST"])
def header_norm_apply():
    try:
        body = request.get_json(force=True, silent=True) or {}
        payload, status = execute_operation_kernel(
            session_id=body.get("sessionId"),
            operation="header_norm_apply",
            api_key=body.get("apiKey"),
            input_data={"decisions": body.get("decisions")},
            options={"mode": "pipeline", "autoPrepare": True, "persist": True},
            request_id=request.headers.get("X-Request-Id"),
        )
        if status != 200:
            return jsonify({"error": payload.get("error"), "missing_requirements": payload.get("missing_requirements")}), status
        return jsonify(payload.get("result") or {})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@header_normalisation_bp.route("/header-norm-preview", methods=["POST"])
def header_norm_preview():
    try:
        body = request.get_json(force=True, silent=True) or {}
        session_id = body.get("sessionId")
        table_key = body.get("tableKey")
        limit = body.get("limit", 100)
        if not session_id or not table_key:
            return jsonify({"error": "Missing sessionId or tableKey"}), 400
        conn = get_session_db(session_id)
        with _session_lock(session_id):
            result = get_table_preview(conn, table_key, int(limit))
        return jsonify(result)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Group-level preview for header normalisation (post-append)
# ---------------------------------------------------------------------------

@header_normalisation_bp.route("/header-norm-group-preview", methods=["POST"])
def header_norm_group_preview():
    """Return first 50 rows per group from appended__ tables."""
    try:
        body = request.get_json(force=True, silent=True) or {}
        session_id = body.get("sessionId")
        group_ids = body.get("groupIds") or []
        limit = int(body.get("limit", 50))
        if not session_id:
            return jsonify({"error": "Missing sessionId"}), 400
        conn = get_session_db(session_id)
        with _session_lock(session_id):
            previews = []
            for gid in group_ids:
                sql = lookup_sql_name(conn, str(gid))
                if not sql or not table_exists(conn, sql):
                    previews.append({"group_id": gid, "columns": [], "rows": [], "total_rows": 0})
                    continue
                cols = read_table_columns(conn, sql)
                rows = read_table(conn, sql, limit)
                previews.append({"group_id": gid, "columns": cols, "rows": rows, "total_rows": table_row_count(conn, sql)})
        return jsonify({"previews": previews})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Excel download / upload for header normalisation
# ---------------------------------------------------------------------------

@header_normalisation_bp.route("/header-norm-download-excel", methods=["POST"])
def header_norm_download_excel():
    """Generate an Excel file for a group with normalization rows above data."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        body = request.get_json(force=True, silent=True) or {}
        session_id = body.get("sessionId")
        group_id = body.get("groupId")
        decisions = body.get("decisions") or []
        if not session_id or not group_id:
            return jsonify({"error": "Missing sessionId or groupId"}), 400

        conn = get_session_db(session_id)
        with _session_lock(session_id):
            sql = lookup_sql_name(conn, str(group_id))
            if not sql or not table_exists(conn, sql):
                return jsonify({"error": f"Table not found for group: {group_id}"}), 404

            cols = read_table_columns(conn, sql)
            rows = read_table(conn, sql, 50)

        decisions_map = {str(d.get("source_col", "")): d for d in decisions}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(group_id)[:31]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        action_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        target_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        meta_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for ci, col in enumerate(cols, 1):
            dec = decisions_map.get(col, {})
            action_cell = ws.cell(row=1, column=ci, value=str(dec.get("action", "KEEP")))
            action_cell.fill = action_fill
            action_cell.font = meta_font
            action_cell.border = thin_border

            target_cell = ws.cell(row=2, column=ci, value=str(dec.get("mapped_to") or dec.get("suggested_std_field") or ""))
            target_cell.fill = target_fill
            target_cell.font = meta_font
            target_cell.border = thin_border

            header_cell = ws.cell(row=3, column=ci, value=col)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal="center")
            header_cell.border = thin_border

        for ri, row in enumerate(rows, 4):
            for ci, col in enumerate(cols, 1):
                val = row[col] if isinstance(row, dict) else (row[ci - 1] if ci - 1 < len(row) else None)
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border

        for ci in range(1, len(cols) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18

        ws.cell(row=1, column=len(cols) + 2, value="ROW 1 = Action (KEEP/AUTO/REVIEW/DROP)")
        ws.cell(row=2, column=len(cols) + 2, value="ROW 2 = Mapped Standard Header")
        ws.cell(row=3, column=len(cols) + 2, value="ROW 3 = Original Column Header")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"{group_id}_header_norm.xlsx"
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except ImportError:
        return jsonify({"error": "openpyxl not installed on server"}), 500
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@header_normalisation_bp.route("/header-norm-upload-excel", methods=["POST"])
def header_norm_upload_excel():
    """Parse an uploaded Excel file and extract normalization decisions from rows 1-3."""
    try:
        import openpyxl

        session_id = request.form.get("sessionId")
        if not session_id:
            return jsonify({"error": "Missing sessionId"}), 400

        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No file uploaded"}), 400

        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        results = {}
        for ws in wb.worksheets:
            group_id = ws.title
            max_col = ws.max_column or 0
            if max_col == 0:
                continue
            actions_row = [str(ws.cell(row=1, column=c).value or "KEEP").strip().upper() for c in range(1, max_col + 1)]
            targets_row = [str(ws.cell(row=2, column=c).value or "").strip() for c in range(1, max_col + 1)]
            headers_row = [str(ws.cell(row=3, column=c).value or "").strip() for c in range(1, max_col + 1)]

            col_decisions = []
            for ci in range(max_col):
                src = headers_row[ci]
                if not src:
                    continue
                action = actions_row[ci] if ci < len(actions_row) else "KEEP"
                if action not in ("AUTO", "REVIEW", "KEEP", "DROP"):
                    action = "KEEP"
                mapped = targets_row[ci] if ci < len(targets_row) else ""
                col_decisions.append({
                    "source_col": src,
                    "action": action,
                    "mapped_to": mapped if mapped else None,
                    "suggested_std_field": mapped if mapped else None,
                    "confidence": 1.0 if mapped else 0.0,
                    "reason": "Imported from Excel",
                    "top_alternatives": [],
                })
            if col_decisions:
                results[group_id] = col_decisions

        return jsonify({"decisions": results})
    except ImportError:
        return jsonify({"error": "openpyxl not installed on server"}), 500
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
