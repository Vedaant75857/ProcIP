"""General normalisation API: upload, headers, preview, download, progress."""

from __future__ import annotations

import csv
import io
import random
import time
import zipfile
from typing import Any, Iterator

import openpyxl
from flask import Blueprint, Response, jsonify, request

from shared.db import (
    all_registered_tables,
    get_session_db,
    lookup_sql_name,
    read_table_columns,
    register_table,
    safe_table_name,
    set_meta,
    store_table_streaming,
    table_row_count,
)
from shared.utils import make_unique

from state.app_state import app_state, get_state

normalisation_bp = Blueprint("normalisation", __name__)


def _clean_header(cells: list) -> list[str]:
    result = []
    for i, c in enumerate(cells):
        if c is None or str(c).strip() == "":
            result.append(f"COLUMN_{i + 1}")
        else:
            result.append(str(c).strip().upper())
    return make_unique(result)


def _clean_cell(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip().upper()
    return s if s else None


def _is_empty_row(vals: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in vals)


def _row_tuple_gen(header: list[str], data_rows: list[list]) -> Iterator[list]:
    n = len(header)
    for r in data_rows:
        if _is_empty_row(r):
            continue
        yield [_clean_cell(r[i]) if i < len(r) else None for i in range(n)]


def _load_excel_sheet(
    conn,
    table_key: str,
    data: bytes,
    sheet_name: str,
    raw_arrays: dict[str, list],
    warnings: list[dict],
) -> None:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        header_raw = None
        raw_arr: list[list] = []
        for row in rows_iter:
            raw_arr.append(list(row))
            if header_raw is None:
                header_raw = list(row)
                break
        for row in rows_iter:
            raw_arr.append(list(row))

        raw_arrays[table_key] = raw_arr[:500]

        if header_raw is None:
            wb.close()
            return

        header = _clean_header(header_raw)
        data_rows = raw_arr[1:]

        def gen():
            yield from _row_tuple_gen(header, data_rows)

        tbl_name = safe_table_name("tbl", table_key)
        store_table_streaming(conn, tbl_name, header, gen())
        register_table(conn, table_key, tbl_name)
        wb.close()
    except Exception as e:
        warnings.append({"file": table_key, "message": str(e)})


def _load_csv(
    conn,
    table_key: str,
    csv_bytes: bytes,
    raw_arrays: dict[str, list],
    warnings: list[dict],
) -> None:
    try:
        text = csv_bytes.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        raw_arr: list[list] = []
        header_raw = None
        for row in reader:
            raw_arr.append(row)
            if header_raw is None:
                header_raw = row
                break
        for row in reader:
            raw_arr.append(row)

        raw_arrays[table_key] = raw_arr[:500]

        if header_raw is None:
            return

        header = _clean_header(header_raw)
        data_rows = raw_arr[1:]

        def gen():
            yield from _row_tuple_gen(header, data_rows)

        tbl_name = safe_table_name("tbl", table_key)
        store_table_streaming(conn, tbl_name, header, gen())
        register_table(conn, table_key, tbl_name)
    except Exception as e:
        warnings.append({"file": table_key, "message": str(e)})


def _load_zip_to_session(conn, file_data: bytes) -> tuple[dict[str, list], list[dict]]:
    raw_arrays: dict[str, list] = {}
    warnings: list[dict] = []

    with zipfile.ZipFile(io.BytesIO(file_data)) as zf:
        for entry in zf.infolist():
            if entry.is_dir():
                continue
            name = entry.filename
            lower = name.lower()

            if lower.endswith((".xlsx", ".xlsm")):
                data = zf.read(name)
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                for sheet in wb.sheetnames:
                    key = f"{name}::{sheet}"
                    _load_excel_sheet(conn, key, data, sheet, raw_arrays, warnings)
                wb.close()
            elif lower.endswith(".csv"):
                key = f"{name}::"
                _load_csv(conn, key, zf.read(name), raw_arrays, warnings)

    for key, arr in raw_arrays.items():
        set_meta(conn, f"rawArray__{key}", arr)

    return raw_arrays, warnings


def _load_single_file(conn, filename: str, file_data: bytes) -> tuple[dict[str, list], list[dict]]:
    raw_arrays: dict[str, list] = {}
    warnings: list[dict] = []
    lower = (filename or "").lower()

    if lower.endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        for sheet in sheets:
            key = f"{filename}::{sheet}"
            _load_excel_sheet(conn, key, file_data, sheet, raw_arrays, warnings)
    elif lower.endswith(".csv"):
        key = f"{filename}::"
        _load_csv(conn, key, file_data, raw_arrays, warnings)
    elif lower.endswith(".zip"):
        return _load_zip_to_session(conn, file_data)
    else:
        warnings.append({"file": filename, "message": "Unsupported file type."})

    for key, arr in raw_arrays.items():
        set_meta(conn, f"rawArray__{key}", arr)

    return raw_arrays, warnings


@normalisation_bp.route("/upload", methods=["POST"])
def upload():
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No file uploaded."}), 400

        file_data = f.read()
        session_id = request.form.get("sessionId") or (
            str(int(time.time() * 1000)) + hex(random.getrandbits(32))[2:]
        )
        conn = get_session_db(session_id)
        raw_arrays, warnings = _load_single_file(conn, f.filename or "upload", file_data)

        tables_out = []
        for row in all_registered_tables(conn):
            rc = table_row_count(conn, row["sql_name"])
            tables_out.append(
                {"tableKey": row["table_key"], "sqlName": row["sql_name"], "rowCount": rc}
            )

        app_state["session_id"] = session_id
        app_state["file_name"] = f.filename
        if tables_out:
            set_meta(conn, "module2_default_table_key", tables_out[0]["tableKey"])

        return jsonify(
            {
                "sessionId": session_id,
                "tables": tables_out,
                "warnings": warnings,
                "rawPreviewKeys": list(raw_arrays.keys()),
            }
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@normalisation_bp.route("/set-api-key", methods=["POST"])
def set_api_key():
    try:
        body = request.get_json(silent=True) or {}
        key = body.get("apiKey") or body.get("openai_api_key")
        if not key:
            return jsonify({"error": "apiKey is required."}), 400
        app_state["openai_api_key"] = key
        return jsonify({"ok": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


def _resolve_table_key(conn, explicit: str | None) -> str:
    if explicit:
        if not lookup_sql_name(conn, explicit):
            raise ValueError("Unknown tableKey.")
        return explicit
    meta_key = None
    try:
        from shared.db import get_meta

        meta_key = get_meta(conn, "module2_default_table_key")
    except Exception:
        pass
    if meta_key and lookup_sql_name(conn, meta_key):
        return meta_key
    reg = all_registered_tables(conn)
    if not reg:
        raise ValueError("No tables loaded for this session.")
    return reg[0]["table_key"]


@normalisation_bp.route("/headers", methods=["GET"])
def headers():
    try:
        sid = app_state.get("session_id")
        if not sid:
            return jsonify({"error": "No active session. Upload a file first."}), 400
        conn = get_session_db(sid)
        table_key = request.args.get("tableKey")
        tk = _resolve_table_key(conn, table_key)
        sql_name = lookup_sql_name(conn, tk)
        assert sql_name
        cols = read_table_columns(conn, sql_name)
        return jsonify({"tableKey": tk, "columns": cols})
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@normalisation_bp.route("/preview", methods=["POST"])
def preview():
    try:
        sid = app_state.get("session_id")
        if not sid:
            return jsonify({"error": "No active session."}), 400
        body = request.get_json(silent=True) or {}
        conn = get_session_db(sid)
        tk = _resolve_table_key(conn, body.get("tableKey"))
        sql_name = lookup_sql_name(conn, tk)
        assert sql_name
        limit = int(body.get("limit", 50))
        from shared.db import read_table

        rows = read_table(conn, sql_name, limit=limit)
        return jsonify({"tableKey": tk, "rows": rows, "rowCount": table_row_count(conn, sql_name)})
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@normalisation_bp.route("/download", methods=["POST"])
def download():
    try:
        sid = app_state.get("session_id")
        if not sid:
            return jsonify({"error": "No active session."}), 400
        body = request.get_json(silent=True) or {}
        conn = get_session_db(sid)
        tk = _resolve_table_key(conn, body.get("tableKey"))
        sql_name = lookup_sql_name(conn, tk)
        assert sql_name
        cols = read_table_columns(conn, sql_name)
        if not cols:
            return jsonify({"error": "Table has no columns."}), 400

        from shared.db import iterate_table

        def generate():
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(cols)
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)
            for row in iterate_table(conn, sql_name):
                w.writerow([row.get(c, "") for c in cols])
                yield buf.getvalue()
                buf.seek(0)
                buf.truncate(0)

        fname = f"{tk.replace('::', '_').replace('/', '_')[:80]}.csv"
        return Response(
            generate(),
            mimetype="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@normalisation_bp.route("/progress", methods=["GET"])
def progress():
    return jsonify(get_state()["progress"])
